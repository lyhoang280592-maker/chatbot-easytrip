import os
import re
import shutil
import openpyxl
import docx
from batch_generate_by_accounting_date import build_contract_docx

# Đường dẫn thư mục
ORIG_DOCX_DIR = os.path.join("HD_Khach_le", "HĐ_Khách lẻ_original")
PASSPORTS_SRC_DIR = os.path.join("downloads", "passports")

TARGET_CONTRACTS_DIR_1 = "output_docx_contracts"
TARGET_CONTRACTS_DIR_2 = os.path.join("HD_Khach_le", "HĐ_Khách lẻ")

TARGET_PASSPORTS_DIR_1 = "output_ho_chieu_khach_hang"
TARGET_PASSPORTS_DIR_2 = os.path.join("HD_Khach_le", "Hộ chiếu khách hàng")

EXCEL_PATH = "danh_sach_khach_hang_theo_accounting_date_01_08_den_19_08.xlsx"


def clean_str(s):
    return re.sub(r'[^a-zA-Z0-9]', '', str(s)).lower()


def organize():
    # 1. Tạo mới / làm sạch các thư mục đích
    for d in [TARGET_CONTRACTS_DIR_1, TARGET_CONTRACTS_DIR_2, TARGET_PASSPORTS_DIR_1, TARGET_PASSPORTS_DIR_2]:
        if os.path.exists(d):
            shutil.rmtree(d)
        os.makedirs(d, exist_ok=True)

    # 2. Đọc danh sách file gốc từ HĐ_Khách lẻ_original
    orig_files = [f for f in os.listdir(ORIG_DOCX_DIR) if f.endswith('.docx')]
    orig_dict = {}
    for f in orig_files:
        name_part = re.sub(r'^\d+_Hop_Dong_', '', f).replace('.docx', '')
        name_part_clean = re.sub(r'_(kocochuky|chuacochuky|chukybikhuat|cancel).*', '', name_part, flags=re.I)
        orig_dict[clean_str(name_part_clean)] = (f, name_part)

    # 3. Đọc danh sách file hộ chiếu từ downloads/passports
    pass_files = os.listdir(PASSPORTS_SRC_DIR)

    # 4. Đọc file Excel 91 khách hàng theo Accounting Date
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    matched_orig_count = 0
    generated_new_count = 0
    passport_copied_count = 0

    print("🚀 Bắt đầu tổ chức lại Hợp đồng và Hộ chiếu...")

    for r in range(5, ws.max_row + 1):
        stt = ws.cell(row=r, column=1).value
        name = str(ws.cell(row=r, column=2).value or '').strip()
        passport_no = str(ws.cell(row=r, column=3).value or '').strip()
        nationality = str(ws.cell(row=r, column=4).value or '').strip()
        acc_date = str(ws.cell(row=r, column=5).value or '').strip()
        state_fee = int(ws.cell(row=r, column=12).value or 662500)
        service_fee = int(ws.cell(row=r, column=13).value or 757500)
        total_amount = int(ws.cell(row=r, column=14).value or (state_fee + service_fee))

        safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', name)
        c_name = clean_str(name)

        # --- A. XỬ LÝ FILE HỘ CHIẾU THEO SỐ THỨ TỰ HỢP ĐỒNG ---
        match_pass_file = None
        if passport_no and passport_no != 'Chưa có':
            for pf in pass_files:
                if passport_no in pf:
                    match_pass_file = pf
                    break
        if not match_pass_file:
            for pf in pass_files:
                if c_name in clean_str(pf):
                    match_pass_file = pf
                    break

        if match_pass_file:
            ext = os.path.splitext(match_pass_file)[1] or '.jpg'
            new_pass_name = f"{stt:02d}_Ho_Chieu_{safe_name}_{passport_no}{ext}"
            src_pass_path = os.path.join(PASSPORTS_SRC_DIR, match_pass_file)
            shutil.copyfile(src_pass_path, os.path.join(TARGET_PASSPORTS_DIR_1, new_pass_name))
            shutil.copyfile(src_pass_path, os.path.join(TARGET_PASSPORTS_DIR_2, new_pass_name))
            passport_copied_count += 1
        else:
            print(f"⚠️ Không tìm thấy ảnh hộ chiếu cho STT {stt:02d}: {name}")

        # --- B. XỬ LÝ FILE HỢP ĐỒNG ---
        # Kiểm tra xem khách có trong file hợp đồng cũ (đã có chữ ký trước đó) không
        found_orig = orig_dict.get(c_name)
        if not found_orig:
            for k, v in orig_dict.items():
                if k in c_name or c_name in k:
                    found_orig = v
                    break

        if found_orig:
            # GIỮ NGUYÊN FILE HỢP ĐỒNG CŨ ĐÃ CÓ CHỮ KÝ CỦA BẠN
            orig_filename, orig_name_part = found_orig
            src_orig_path = os.path.join(ORIG_DOCX_DIR, orig_filename)
            new_contract_filename = f"{stt:02d}_Hop_Dong_{orig_name_part}.docx"
            
            dst_1 = os.path.join(TARGET_CONTRACTS_DIR_1, new_contract_filename)
            dst_2 = os.path.join(TARGET_CONTRACTS_DIR_2, new_contract_filename)
            
            shutil.copyfile(src_orig_path, dst_1)
            shutil.copyfile(src_orig_path, dst_2)
            matched_orig_count += 1
        else:
            # TẠO FILE HỢP ĐỒNG MỚI (KHÔNG TRÍCH XUẤT CHỮ KÝ, ĐỂ TRỐNG Ô KÝ)
            new_contract_filename = f"{stt:02d}_Hop_Dong_{safe_name}.docx"
            dst_1 = os.path.join(TARGET_CONTRACTS_DIR_1, new_contract_filename)
            dst_2 = os.path.join(TARGET_CONTRACTS_DIR_2, new_contract_filename)

            date_vi = f"ngày {acc_date.split('/')[0]} tháng {acc_date.split('/')[1]} năm {acc_date.split('/')[2]}" if '/' in acc_date else "ngày 16 tháng 08 năm 2026"
            date_en = "August 16, 2026"
            if '/' in acc_date:
                try:
                    from datetime import datetime
                    dt = datetime.strptime(acc_date, "%d/%m/%Y")
                    date_en = dt.strftime("%B %d, %Y")
                except:
                    pass

            contract_data = {
                "contract_no": f"{stt:03d}",
                "date_vi": date_vi,
                "date_en": date_en,
                "customer_name": name.upper(),
                "passport_no": passport_no,
                "date_of_issue": "20/05/2022",
                "nationality": nationality,
                "service_fee": service_fee,
                "state_fee": state_fee,
                "signature_image_path": None # KHÔNG TRÍCH XUẤT CHỮ KÝ TỰ ĐỘNG
            }

            build_contract_docx(dst_1, contract_data)
            shutil.copyfile(dst_1, dst_2)
            generated_new_count += 1

    print(f"\n🎉 HOÀN TẤT TỔ CHỨC:")
    print(f"1. Tổng số hợp đồng: {matched_orig_count + generated_new_count} file")
    print(f"   - Giữ nguyên file gốc đã có chữ ký: {matched_orig_count} file")
    print(f"   - Tạo mới không chèn chữ ký tự động: {generated_new_count} file")
    print(f"2. Tổng số file hộ chiếu đã sắp xếp theo STT: {passport_copied_count} file")


if __name__ == "__main__":
    organize()
