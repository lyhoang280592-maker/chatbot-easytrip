import os
import re
import json
import asyncio
import httpx
from datetime import datetime
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from generate_contracts import extract_and_clean_signature

load_dotenv("/Users/phamtranthuyvy/Projects/chatbot-easytrip/.env")

LARK_APP_ID = os.getenv("LARK_APP_ID")
LARK_APP_SECRET = os.getenv("LARK_APP_SECRET")
LARK_APP_TOKEN = os.getenv("LARK_APP_TOKEN")
TABLE_ID = "tbluosVi3sQS9gIS"


async def restore_data():
    url = "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal"
    async with httpx.AsyncClient(timeout=15) as client:
        r = await client.post(url, json={"app_id": LARK_APP_ID, "app_secret": LARK_APP_SECRET})
        token = r.json().get("tenant_access_token")

    if not token:
        print("❌ Không lấy được Lark Token!")
        return

    headers = {"Authorization": f"Bearer {token}"}
    all_records = []
    page_token = None
    while True:
        p_url = f"https://open.larksuite.com/open-apis/bitable/v1/apps/{LARK_APP_TOKEN}/tables/{TABLE_ID}/records?page_size=500"
        if page_token: p_url += f"&page_token={page_token}"
        async with httpx.AsyncClient(timeout=30) as client:
            r_rec = await client.get(p_url, headers=headers)
            data_res = r_rec.json().get("data", {})
            items = data_res.get("items", [])
            all_records.extend(items)
            if not data_res.get("has_more"): break
            page_token = data_res.get("page_token")

    start_ts = datetime(2026, 8, 1, 0, 0, 0).timestamp() * 1000
    end_ts = datetime(2026, 8, 19, 23, 59, 59).timestamp() * 1000

    matched = []
    for it in all_records:
        f = it.get("fields", {})
        ev_ts = f.get("Event Date")
        if ev_ts and start_ts <= ev_ts <= end_ts:
            code_ev = f.get("Code EV")
            ev_file = f.get("EV")
            if code_ev or ev_file:
                matched.append(f)

    # Sắp xếp theo Event Date giảm dần
    matched.sort(key=lambda x: x.get("Event Date") or 0, reverse=True)
    print(f"📊 Tìm thấy {len(matched)} khách hàng có E-visa từ 01/08 đến 19/08.")

    # 1. TẠO FILE EXCEL DANH SÁCH
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title="Khach_Hang_Co_Evisa")
    else:
        ws.title = "Khach_Hang_Co_Evisa"

    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "DANH SÁCH KHÁCH HÀNG CÓ E-VISA (01/08/2026 - 19/08/2026)"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:J2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Tổng cộng: {len(matched)} khách hàng | Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub_cell.font = Font(name="Arial", size=10, italic=True, color="333333")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    excel_headers = [
        "STT", "Ngày Event", "Ngày nộp (Apply)", "Tên khách hàng", "Quốc tịch",
        "Loại dịch vụ", "Mã E-Visa (Code EV)", "Tệp đính kèm EV", "Doanh thu (VNĐ)", "Trạng thái thanh toán"
    ]

    header_row = 4
    ws.row_dimensions[header_row].height = 28
    for col_idx, h in enumerate(excel_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.border = Border(top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"),
                             left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"))

    thin_border = Border(top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"),
                         left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"))

    total_revenue = 0

    for idx, c in enumerate(matched, 1):
        row_idx = header_row + idx
        ws.row_dimensions[row_idx].height = 22

        ev_dt = datetime.fromtimestamp(c.get("Event Date") / 1000).strftime("%d/%m/%Y") if c.get("Event Date") else ""
        app_dt = datetime.fromtimestamp(c.get("Apply Date") / 1000).strftime("%d/%m/%Y") if c.get("Apply Date") else ""
        name = str(c.get("Tên khách (Name)", "") or "").strip().upper()
        nat = c.get("Quốc tịch (National)", [""])[0] if c.get("Quốc tịch (National)") else ""
        srv = c.get("Loại dịch vụ (type)", [""])[0] if c.get("Loại dịch vụ (type)") else ""
        code_ev = str(c.get("Code EV", "") or "").strip()
        
        ev_files = c.get("EV", [])
        ev_file_name = ev_files[0].get("name", "Có tệp") if isinstance(ev_files, list) and len(ev_files) > 0 else ("Không có tệp" if not ev_files else "Có tệp")
        
        try:
            rev_val = int(str(c.get("Sales revenue", "0") or "0").replace(".", "").replace(",", ""))
        except (ValueError, TypeError):
            rev_val = 0
        total_revenue += rev_val

        status = c.get("Transaction Status", [""])[0] if c.get("Transaction Status") else ""

        row_values = [idx, ev_dt, app_dt, name, nat, srv, code_ev, ev_file_name, rev_val, status]

        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = thin_border
            
            if col_idx in [1, 2, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [4, 7]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if col_idx == 4: cell.font = Font(name="Arial", size=9, bold=True)
            elif col_idx in [5, 6, 8, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 9:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"

            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    total_row = header_row + len(matched) + 1
    ws.row_dimensions[total_row].height = 26
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
    sum_label = ws.cell(row=total_row, column=1, value="TỔNG DOANH THU CÁC KHÁCH CÓ E-VISA")
    sum_label.font = Font(name="Arial", size=10, bold=True)
    sum_label.alignment = Alignment(horizontal="right", vertical="center")
    
    sum_val = ws.cell(row=total_row, column=9, value=total_revenue)
    sum_val.font = Font(name="Arial", size=10, bold=True, color="C00000")
    sum_val.alignment = Alignment(horizontal="right", vertical="center")
    sum_val.number_format = "#,##0"

    for col in range(1, 11):
        c_cell = ws.cell(row=total_row, column=col)
        c_cell.border = Border(top=Side(style="thin", color="000000"), bottom=Side(style="double", color="000000"))
        c_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_idx = col[0].column
        if col_idx is not None:
            col_letter = get_column_letter(int(col_idx))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    excel_path = "danh_sach_khach_hang_co_evisa_01_08_den_19_08.xlsx"
    wb.save(excel_path)
    print(f"✅ Đã khôi phục file Excel: {excel_path}")

    # 2. TẢI ẢNH HỘ CHIẾU VÀ TRÍCH XUẤT CHỮ KÝ
    os.makedirs("downloads/passports", exist_ok=True)
    os.makedirs("extracted_signatures", exist_ok=True)

    print(f"📥 Đang tải ảnh hộ chiếu và trích xuất chữ ký cho {len(matched)} khách hàng...")

    async with httpx.AsyncClient(timeout=30) as client:
        for idx, c in enumerate(matched, 1):
            name = str(c.get("Tên khách (Name)", "CUSTOMER") or "CUSTOMER").strip()
            code_ev = str(c.get("Code EV", "") or "").strip()
            safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', name.strip())
            passport_match = re.search(r'[A-Za-z0-9]{7,9}$', code_ev)
            passport_no = passport_match.group(0) if passport_match else (code_ev.replace("E26", "")[:9] if code_ev else "767587433")

            sig_path = os.path.join("extracted_signatures", f"{safe_name}_{passport_no}_sig.png")
            passports = c.get("Ảnh hộ chiếu")
            if passports and isinstance(passports, list) and len(passports) > 0:
                p_item = passports[0]
                ft = p_item.get("file_token")
                p_path = os.path.join("downloads/passports", f"{safe_name}_{passport_no}.jpg")
                if not os.path.exists(p_path) and ft:
                    try:
                        dl_url = f"https://open.larksuite.com/open-apis/drive/v1/medias/{ft}/download"
                        r_dl = await client.get(dl_url, headers=headers)
                        if r_dl.status_code == 200:
                            with open(p_path, "wb") as img_f:
                                img_f.write(r_dl.content)
                    except Exception as e:
                        print(f"Lỗi tải ảnh {name}: {e}")
                
                if os.path.exists(p_path):
                    extract_and_clean_signature(p_path, sig_path)

    print("🎉 Khôi phục hoàn tất toàn bộ danh sách, ảnh hộ chiếu và chữ ký!")


if __name__ == "__main__":
    asyncio.run(restore_data())
