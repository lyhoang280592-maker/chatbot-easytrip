import os
import re
import asyncio
import httpx
from datetime import datetime
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from generate_contracts import build_contract_pdf

load_dotenv("/Users/phamtranthuyvy/Projects/chatbot-easytrip/.env")

LARK_APP_ID = os.getenv("LARK_APP_ID")
LARK_APP_SECRET = os.getenv("LARK_APP_SECRET")
LARK_APP_TOKEN = os.getenv("LARK_APP_TOKEN")
TABLE_ID = "tbluosVi3sQS9gIS" # Data KH 2026

async def get_tenant_access_token():
    url = "https://open.larksuite.com/open-apis/auth/v3/tenant_access_token/internal"
    async with httpx.AsyncClient(timeout=15) as client:
        r = await client.post(url, json={"app_id": LARK_APP_ID, "app_secret": LARK_APP_SECRET})
        return r.json().get("tenant_access_token")

async def fetch_customers_by_event_date(start_date_str="01/08/2026", end_date_str=None, only_reconciled=False):
    """
    Tải và lọc danh sách khách hàng có Event Date từ start_date_str đến end_date_str (mặc định hôm nay).
    """
    token = await get_tenant_access_token()
    if not token:
        print("❌ Không lấy được Lark Tenant Access Token!")
        return []

    headers = {"Authorization": f"Bearer {token}"}
    print(f"📡 Đang kết nối Lark Base để tải hồ sơ khách hàng...")
    
    all_records = []
    page_token = None
    while True:
        p_url = f"https://open.larksuite.com/open-apis/bitable/v1/apps/{LARK_APP_TOKEN}/tables/{TABLE_ID}/records?page_size=500"
        if page_token: p_url += f"&page_token={page_token}"
        async with httpx.AsyncClient(timeout=30) as client:
            r_rec = await client.get(p_url, headers=headers)
            data = r_rec.json().get("data", {})
            items = data.get("items", [])
            all_records.extend(items)
            if not data.get("has_more"): break
            page_token = data.get("page_token")

    # Xử lý mốc thời gian lọc
    if start_date_str:
        dt_start = datetime.strptime(start_date_str, "%d/%m/%Y")
        start_ts = dt_start.timestamp() * 1000
    else:
        start_ts = 0

    if end_date_str:
        dt_end = datetime.strptime(end_date_str, "%d/%m/%Y").replace(hour=23, minute=59, second=59)
        end_ts = dt_end.timestamp() * 1000
    else:
        end_ts = datetime.now().replace(hour=23, minute=59, second=59).timestamp() * 1000

    matched_customers = []
    for it in all_records:
        f = it.get("fields", {})
        ev_ts = f.get("Event Date")
        
        # Kiểm tra điều kiện Event Date nằm trong khoảng
        if ev_ts and start_ts <= ev_ts <= end_ts:
            if only_reconciled:
                passports = f.get("Ảnh hộ chiếu")
                invoices = f.get("Paid Invoice")
                v_vn = f.get("Visa VN")
                if passports and isinstance(passports, list) and len(passports) > 0 and invoices and v_vn:
                    matched_customers.append(f)
            else:
                matched_customers.append(f)

    # Sắp xếp theo ngày Event Date giảm dần (mới nhất lên đầu)
    matched_customers.sort(key=lambda x: x.get("Event Date") or 0, reverse=True)
    print(f"✅ Đã lọc thành công {len(matched_customers)} khách hàng có Event Date từ {start_date_str} đến hiện tại.")
    return matched_customers

async def fetch_reconciled_customers():
    return await fetch_customers_by_event_date("01/08/2026", None, only_reconciled=True)

def export_to_excel(customers, output_file="danh_sach_doi_soat_khach_hang.xlsx"):
    """Xuất danh sách khách hàng đã đối soát ra tệp Excel định dạng chuẩn đẹp"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Doi_Soat_KH_2026"

    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    headers = [
        "STT", "Ngày đi / Nộp", "Họ và Tên", "Quốc tịch", "Loại dịch vụ", 
        "Lệ phí Nhà nước (Visa VN)", "Phí dịch vụ tư vấn", "Tổng doanh thu", 
        "Trạng thái thanh toán", "Kênh nguồn", "Mã hồ sơ EV"
    ]

    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    for i, c in enumerate(customers, 1):
        name = c.get("Tên khách (Name)", "")
        nat = c.get("Quốc tịch (National)", [""])[0] if c.get("Quốc tịch (National)") else ""
        serv = c.get("Loại dịch vụ (type)", [""])[0] if c.get("Loại dịch vụ (type)") else ""
        code_ev = str(c.get("Code EV", "") or "")
        status = c.get("Transaction Status", [""])[0] if c.get("Transaction Status") else ""
        channel = c.get("Nguồn( Channel)", [""])[0] if c.get("Nguồn( Channel)") else ""
        
        # Tiền tệ
        try:
            state_fee = int(str(c.get("Visa VN", "0")).replace(".", "").replace(",", ""))
        except: state_fee = 0
        
        try:
            total_rev = int(str(c.get("Sales revenue", "0")).replace(".", "").replace(",", ""))
        except: total_rev = 0
        
        service_fee = max(0, total_rev - state_fee)

        ev_ts = c.get("Event Date") or c.get("Apply Date")
        d_str = datetime.fromtimestamp(ev_ts/1000).strftime("%d/%m/%Y") if ev_ts else ""

        row = [
            i, d_str, name, nat, serv, 
            state_fee, service_fee, total_rev, 
            status, channel, code_ev
        ]
        ws.append(row)

        r_idx = i + 1
        ws.row_dimensions[r_idx].height = 20
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_num)
            cell.font = cell_font
            cell.border = thin_border
            if col_num in [1, 2]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [6, 7, 8]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Thêm dòng Tổng Cộng ở cuối
    last_row = len(customers) + 2
    ws.cell(row=last_row, column=3, value="TỔNG CỘNG").font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=last_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num in [6, 7, 8]:
        col_letter = openpyxl.utils.get_column_letter(col_num)
        cell = ws.cell(row=last_row, column=col_num, value=f"=SUM({col_letter}2:{col_letter}{last_row-1})")
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right", vertical="center")

    for col_num in range(1, len(headers) + 1):
        c = ws.cell(row=last_row, column=col_num)
        c.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
        c.fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")

    # Tự động căn chỉnh độ rộng cột
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 13)

    wb.save(output_file)
    print(f"📊 Đã xuất file Excel đối soát thành công: {output_file}")
    return output_file

from generate_contracts import build_contract_pdf, extract_and_clean_signature

async def download_and_extract_signatures(customers, sig_dir="extracted_signatures", passport_dir="downloads/passports"):
    """
    Tải ảnh hộ chiếu từ Lark Base và tự động trích xuất chữ ký trong suốt cho từng khách hàng
    """
    os.makedirs(sig_dir, exist_ok=True)
    os.makedirs(passport_dir, exist_ok=True)
    
    token = await get_tenant_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    
    print(f"📥 Đang tải ảnh hộ chiếu và trích xuất chữ ký cho {len(customers)} khách hàng...")
    success_count = 0
    async with httpx.AsyncClient(timeout=30) as client:
        for idx, c in enumerate(customers, 1):
            passports = c.get("Ảnh hộ chiếu")
            name = c.get("Tên khách (Name)", "UNKNOWN")
            code_ev = str(c.get("Code EV", "") or "")
            passport_no = code_ev.replace("E26", "")[:9] if code_ev else "NO_PASSPORT"
            safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', name.strip())
            
            if passports and isinstance(passports, list) and len(passports) > 0:
                p = passports[0]
                ft = p.get("file_token")
                p_path = os.path.join(passport_dir, f"{safe_name}_{passport_no}.jpg")
                sig_path = os.path.join(sig_dir, f"{safe_name}_{passport_no}_sig.png")
                
                # Tải ảnh nếu chưa có
                if not os.path.exists(p_path) and ft:
                    dl_url = f"https://open.larksuite.com/open-apis/drive/v1/medias/{ft}/download"
                    try:
                        r_dl = await client.get(dl_url, headers=headers)
                        if r_dl.status_code == 200:
                            with open(p_path, "wb") as img_f:
                                img_f.write(r_dl.content)
                    except Exception as e:
                        print(f"⚠️ Lỗi tải ảnh hộ chiếu của {name}: {e}")
                        continue
                
                # Trích xuất chữ ký
                if os.path.exists(p_path):
                    if extract_and_clean_signature(p_path, sig_path):
                        c["signature_image_path"] = sig_path
                        success_count += 1
                        print(f"[{idx}/{len(customers)}] ✍️ Trích xuất chữ ký thành công: {name}")

    print(f"✨ Đã trích xuất thành công {success_count}/{len(customers)} chữ ký khách hàng.")
    return customers

def generate_customer_contract(customer, output_dir="output_contracts"):
    """Tạo file Hợp đồng 7 trang song ngữ chuẩn cho 1 khách hàng cụ thể (kèm chữ ký)"""
    os.makedirs(output_dir, exist_ok=True)
    
    name = customer.get("Tên khách (Name)", "CUSTOMER")
    nat = customer.get("Quốc tịch (National)", ["Nga / Russian"])[0] if customer.get("Quốc tịch (National)") else "Nga / Russian"
    code_ev = str(customer.get("Code EV", "") or "")
    
    try:
        state_fee = int(str(customer.get("Visa VN", "678629")).replace(".", "").replace(",", ""))
    except: state_fee = 678629
    
    try:
        total_rev = int(str(customer.get("Sales revenue", "1420000")).replace(".", "").replace(",", ""))
    except: total_rev = 1420000
    
    service_fee = max(0, total_rev - state_fee)
    if service_fee == 0:
        service_fee = 757500
        total_rev = service_fee + state_fee

    ev_ts = customer.get("Event Date") or customer.get("Apply Date")
    if ev_ts:
        dt = datetime.fromtimestamp(ev_ts / 1000)
        date_vi = dt.strftime("ngày %d tháng %m năm %Y")
        date_en = dt.strftime("%B %d, %Y")
    else:
        date_vi = "ngày 16 tháng 06 năm 2026"
        date_en = "June 16, 2026"

    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', name.strip())
    pdf_path = os.path.join(output_dir, f"Hop_Dong_{safe_name}.pdf")
    passport_no = code_ev.replace("E26", "")[:9] if code_ev else "767587433"

    contract_data = {
        "contract_no": str(abs(hash(name)) % 900 + 100).zfill(3),
        "date_vi": date_vi,
        "date_en": date_en,
        "customer_name": name.upper(),
        "passport_no": passport_no,
        "date_of_issue": "20/05/2022",
        "nationality": nat,
        "service_fee": service_fee,
        "state_fee": state_fee,
        "signature_image_path": customer.get("signature_image_path") # Chữ ký đã trích xuất
    }

    build_contract_pdf(pdf_path, contract_data)
    return pdf_path

async def main():
    # Lọc khách hàng có Event Date từ 01/08/2026 đến hiện tại có đủ Hộ chiếu + Biên lai + Lệ phí
    customers = await fetch_customers_by_event_date("01/08/2026", None, only_reconciled=True)
    if customers:
        # Xuất file Excel đối soát
        export_to_excel(customers, "danh_sach_khach_hang_khop_ho_chieu_bien_lai.xlsx")
        
        # Tải ảnh hộ chiếu và trích xuất chữ ký
        customers_with_sig = await download_and_extract_signatures(customers)

if __name__ == "__main__":
    asyncio.run(main())
