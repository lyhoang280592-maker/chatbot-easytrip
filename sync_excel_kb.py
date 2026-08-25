import openpyxl
import json
import os
import sys

# Reconfigure stdout for UTF-8 on Windows
if hasattr(sys.stdout, "reconfigure"):
    getattr(sys.stdout, "reconfigure")(encoding="utf-8")

def sync():
    base_dir = "C:/Users/Admin/.gemini/antigravity/scratch/backend"
    excel_path = os.path.join(base_dir, "easytrip_qa_master.xlsx")
    json_path = os.path.join(base_dir, "extracted_qa_excel.json")
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel master file not found at '{excel_path}'")
        sys.exit(1)
        
    print(f"Reading Excel Q&A from: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    if "Q&A Master" not in wb.sheetnames:
        print("Error: Sheet 'Q&A Master' not found in the workbook!")
        sys.exit(1)
        
    ws = wb["Q&A Master"]
    qa_list = []
    
    # Iterate rows starting from row 2 (skipping header)
    for row in range(2, ws.max_row + 1):
        cat = ws.cell(row=row, column=2).value
        q = ws.cell(row=row, column=3).value
        a = ws.cell(row=row, column=4).value
        note = ws.cell(row=row, column=5).value
        
        if not q or not a:
            continue
            
        q_str = str(q).strip()
        a_str = str(a).strip()
        
        # If there's an operational warning/note, append it neatly
        if note and str(note).strip():
            a_str += f"\n\n🚨 [LƯU Ý NGHIỆP VỤ]: {str(note).strip()}"
            
        qa_list.append({
            "question": q_str,
            "answer": a_str
        })
        
    print(f"Parsed {len(qa_list)} valid Q&A pairs from Excel.")
    
    # Save to JSON file
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(qa_list, f, ensure_ascii=False, indent=2)
        
    print(f"Successfully saved to: {json_path}")
    
    # Rebuild RAG index
    sys.path.append(base_dir)
    try:
        import knowledge_rag
        knowledge_rag.rebuild()
        print("RAG index rebuilt successfully with the new Excel Q&As!")
    except Exception as e:
        print(f"Warning: Failed to rebuild RAG index programmatically: {e}")

if __name__ == "__main__":
    sync()
